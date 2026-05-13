#!/usr/bin/env python
from __future__ import annotations

import argparse
import hashlib
import json
import os
import pickle
import re
import sys
import time
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import unquote, urljoin, urlparse

import numpy as np
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader
from requests.utils import requote_uri
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer


ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
RAW = DATA / "raw"
RAW_DOCS = RAW / "documents"
EXTRACTED = DATA / "extracted"
PROCESSED = DATA / "processed"
INDEX_URL = "https://www.ecb.europa.eu/paym/target/target-professional-use-documents-links/tips/html/index.en.html"
USER_AGENT = "Mozilla/5.0 (compatible; TIPSLocalBot/1.0; +local-ingestion)"
TEXT_EXTENSIONS = {
    ".txt",
    ".xml",
    ".xsd",
    ".csv",
    ".json",
    ".md",
    ".html",
    ".htm",
    ".yaml",
    ".yml",
}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def log(message: str) -> None:
    print(f"[tips-ingest] {message}", flush=True)


def sha1_text(text: str, n: int = 12) -> str:
    return hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()[:n]


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def safe_slug(text: str, max_len: int = 90) -> str:
    text = unquote(text or "")
    text = text.replace("&", " and ")
    text = re.sub(r"[^A-Za-z0-9._ -]+", " ", text)
    text = re.sub(r"\s+", "-", text.strip())
    text = text.strip("-._")
    if not text:
        text = "document"
    return text[:max_len].strip("-._") or "document"


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"[ \t\r\f\v]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def normalize_url(url: str) -> str:
    return requote_uri(url.strip())


def url_extension(url: str) -> str:
    path = urlparse(url).path
    suffix = Path(unquote(path)).suffix.lower()
    if suffix:
        return suffix
    if "/html/" in path or path.endswith("/"):
        return ".html"
    return ".html"


def previous_header_title(content_box) -> str:
    prev = content_box.previous_sibling
    while prev is not None:
        if getattr(prev, "name", None) == "div" and "header" in (prev.get("class") or []):
            node = prev.select_one(".title")
            return " ".join((node or prev).get_text(" ", strip=True).split())
        prev = prev.previous_sibling
    return ""


def context_for_anchor(anchor, main) -> dict:
    headers: list[str] = []
    for parent in anchor.parents:
        if parent is main:
            break
        if getattr(parent, "name", None) == "div" and "content-box" in (parent.get("class") or []):
            title = previous_header_title(parent)
            if title:
                headers.append(title)
    headers = list(reversed(headers))

    h3_text = ""
    for h3 in anchor.find_all_previous("h3"):
        if h3.find_parent("main") is main:
            h3_text = " ".join(h3.get_text(" ", strip=True).split())
            break

    return {
        "h3": h3_text,
        "accordion": headers,
        "path": [x for x in [h3_text, *headers] if x],
    }


def release_from_text(text: str) -> str:
    text = text or ""
    match = re.search(r"R(20\d{2})[._-]?(NOV|OCT|JUN|MAR)", text, re.I)
    if match:
        return f"R{match.group(1)}.{match.group(2).upper()}"
    match = re.search(r"Release\s+(\d+(?:\.\d+)?)", text, re.I)
    if match:
        return f"Release {match.group(1)}"
    if re.search(r"March\s+2024", text, re.I):
        return "Ad-hoc March 2024"
    return ""


def doc_family(title: str, url: str, context: dict) -> str:
    hay = " ".join([title, url, " ".join(context.get("path", []))]).lower()
    if "production problem" in hay:
        return "production_problems"
    if "pricing" in hay:
        return "pricing"
    if "change request" in hay:
        return "change_requests"
    if "user requirement" in hay or "tips urd" in hay:
        return "tips_urd"
    if "mpl" in hay and ("udfs" in hay or "schema" in hay or "example" in hay):
        return "mpl_udfs"
    if "uhb" in hay or "user handbook" in hay:
        return "tips_uhb"
    if "udfs" in hay or "detailed functional" in hay:
        return "tips_udfs"
    if "mept" in hay or "message exchange" in hay or "connectivity" in hay or "nsp" in hay or "gosign" in hay:
        return "connectivity"
    if "hosting terms" in hay or "legal" in hay or "guideline" in hay:
        return "legal"
    if "training" in hay or "live demo" in hay or "one pager" in hay or "validations" in hay:
        return "training_and_featured_topics"
    if "participation" in hay or "onboarding" in hay or "instant payments" in hay:
        return "participation"
    if "release" in hay or release_from_text(hay):
        return "release_documentation"
    if "consultative group" in hay or "meetdoc" in hay:
        return "consultative_group"
    if "shared documentation" in hay:
        return "shared_features"
    return "general"


def category_for(title: str, url: str, context: dict) -> str:
    family = doc_family(title, url, context)
    release = release_from_text(" ".join([title, url, " ".join(context.get("path", []))]))
    if release:
        return "releases"
    return family


def revision_status(title: str, url: str) -> str:
    hay = f"{title} {url}".lower()
    if "with revisions" in hay or "_rev" in hay or "-rev" in hay:
        return "with_revisions"
    if "clean" in hay:
        return "clean"
    return ""


@dataclass
class Document:
    id: str
    title: str
    url: str
    ext: str
    category: str
    family: str
    release: str = ""
    revision_status: str = ""
    contexts: list[dict] = field(default_factory=list)
    status: str = "pending"
    local_path: str = ""
    source_host: str = ""
    media_skipped: bool = False
    error: str = ""
    sha256: str = ""
    size_bytes: int = 0
    extracted_chars: int = 0
    extracted_units: int = 0
    text_path: str = ""

    def to_dict(self) -> dict:
        return dict(self.__dict__)


def parse_index(html: str, include_media: bool = False) -> tuple[list[Document], list[dict]]:
    soup = BeautifulSoup(html, "lxml")
    main = soup.find("main") or soup
    by_url: dict[str, Document] = {}
    skipped: list[dict] = []

    for anchor in main.find_all("a", href=True):
        title = " ".join(anchor.get_text(" ", strip=True).split())
        href = urljoin(INDEX_URL, anchor["href"])
        href = normalize_url(href)
        parsed = urlparse(href)
        if not title or href.endswith("#") or parsed.scheme in {"mailto", "javascript"}:
            continue

        host = parsed.netloc.lower()
        ext = url_extension(href)
        is_media = ext in {".mp4", ".mov", ".avi"}
        is_swift_login = "login.swift.com" in host
        if is_swift_login:
            skipped.append({"title": title, "url": href, "reason": "login_only_external_reference"})
            continue
        if is_media and not include_media:
            skipped.append({"title": title, "url": href, "reason": "media_not_text_document"})
            continue

        context = context_for_anchor(anchor, main)
        family = doc_family(title, href, context)
        release = release_from_text(" ".join([title, href, " ".join(context.get("path", []))]))
        category = category_for(title, href, context)
        key = href
        if key not in by_url:
            doc_id = sha1_text(key)
            by_url[key] = Document(
                id=doc_id,
                title=title,
                url=href,
                ext=ext,
                category=category,
                family=family,
                release=release,
                revision_status=revision_status(title, href),
                contexts=[context],
                source_host=host,
                media_skipped=is_media and not include_media,
            )
        else:
            doc = by_url[key]
            if context not in doc.contexts:
                doc.contexts.append(context)
            if not doc.release and release:
                doc.release = release
            if doc.family == "general" and family != "general":
                doc.family = family
            if doc.category == "general" and category != "general":
                doc.category = category

    docs = list(by_url.values())
    docs.sort(key=lambda d: (d.category, d.release, d.family, d.title.lower()))
    return docs, skipped


def fetch_index(session: requests.Session, force: bool) -> str:
    RAW.mkdir(parents=True, exist_ok=True)
    target = RAW / "ecb_tips_index.html"
    legacy = ROOT / "ecb_tips_index.raw.html"
    if target.exists() and not force:
        return target.read_text(encoding="utf-8", errors="ignore")
    if legacy.exists() and not force:
        html = legacy.read_text(encoding="utf-8", errors="ignore")
        target.write_text(html, encoding="utf-8")
        return html
    response = session.get(INDEX_URL, timeout=60)
    response.raise_for_status()
    html = response.text
    target.write_text(html, encoding="utf-8")
    return html


def local_path_for(doc: Document) -> Path:
    release = safe_slug(doc.release, 40) if doc.release else ""
    folder = RAW_DOCS / safe_slug(doc.category, 60)
    if release:
        folder = folder / release
    suffix = doc.ext if doc.ext.startswith(".") else f".{doc.ext}"
    name = safe_slug(doc.title, 80)
    return folder / f"{name}__{doc.id}{suffix}"


def download_doc(session: requests.Session, doc: Document, force: bool = False) -> Document:
    target = local_path_for(doc)
    target.parent.mkdir(parents=True, exist_ok=True)
    doc.local_path = str(target.relative_to(ROOT))

    if target.exists() and target.stat().st_size > 0 and not force:
        doc.status = "downloaded"
        doc.size_bytes = target.stat().st_size
        doc.sha256 = sha256_file(target)
        return doc

    try:
        with session.get(doc.url, stream=True, timeout=120, allow_redirects=True) as response:
            response.raise_for_status()
            tmp = target.with_suffix(target.suffix + ".tmp")
            with tmp.open("wb") as fh:
                for block in response.iter_content(chunk_size=1024 * 256):
                    if block:
                        fh.write(block)
            tmp.replace(target)
        doc.status = "downloaded"
        doc.size_bytes = target.stat().st_size
        doc.sha256 = sha256_file(target)
    except Exception as exc:
        doc.status = "failed"
        doc.error = repr(exc)
        log(f"download failed: {doc.title} -> {exc}")
    return doc


def extract_pdf(path: Path) -> list[dict]:
    units: list[dict] = []
    reader = PdfReader(str(path))
    if reader.is_encrypted:
        try:
            reader.decrypt("")
        except Exception:
            pass
    for idx, page in enumerate(reader.pages, start=1):
        try:
            text = clean_text(page.extract_text() or "")
        except Exception as exc:
            text = f"[page extraction failed: {exc!r}]"
        if text:
            units.append({"unit_type": "page", "unit": idx, "text": text})
    return units


def extract_html(path: Path) -> list[dict]:
    raw = path.read_text(encoding="utf-8", errors="ignore")
    soup = BeautifulSoup(raw, "lxml")
    for tag in soup(["script", "style", "noscript", "svg", "header", "footer", "nav"]):
        tag.decompose()
    main = soup.find("main") or soup.body or soup
    title = soup.find("title")
    parts = []
    if title:
        parts.append(" ".join(title.get_text(" ", strip=True).split()))
    for node in main.find_all(["h1", "h2", "h3", "h4", "p", "li", "td", "th"]):
        text = " ".join(node.get_text(" ", strip=True).split())
        if text:
            parts.append(text)
    text = clean_text("\n".join(parts))
    return [{"unit_type": "html", "unit": 1, "text": text}] if text else []


def extract_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    units: list[dict] = []
    for ws in wb.worksheets:
        lines = [f"Sheet: {ws.title}"]
        for row in ws.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if values:
                lines.append(" | ".join(values))
        text = clean_text("\n".join(lines))
        if text:
            units.append({"unit_type": "sheet", "unit": ws.title, "text": text})
    return units


def extract_zip(path: Path) -> list[dict]:
    units: list[dict] = []
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        inventory = "\n".join(names)
        units.append({"unit_type": "zip_inventory", "unit": 1, "text": f"ZIP inventory:\n{inventory}"})
        for name in names:
            suffix = Path(name).suffix.lower()
            info = archive.getinfo(name)
            if info.is_dir() or suffix not in TEXT_EXTENSIONS or info.file_size > 1024 * 1024:
                continue
            try:
                raw = archive.read(name)
                text = raw.decode("utf-8", errors="ignore")
                text = clean_text(text)
            except Exception:
                continue
            if text:
                units.append({"unit_type": "zip_member", "unit": name, "text": f"ZIP member: {name}\n{text}"})
    return units


def extract_text_for_doc(doc: Document) -> tuple[Document, list[dict]]:
    path = ROOT / doc.local_path
    units: list[dict] = []
    try:
        if doc.ext == ".pdf":
            units = extract_pdf(path)
        elif doc.ext in {".html", ".htm", ".rss"}:
            units = extract_html(path)
        elif doc.ext == ".xlsx":
            units = extract_xlsx(path)
        elif doc.ext == ".zip":
            units = extract_zip(path)
        else:
            try:
                units = [{"unit_type": "file", "unit": 1, "text": clean_text(path.read_text(encoding="utf-8", errors="ignore"))}]
            except UnicodeDecodeError:
                units = []
        text = "\n\n".join(f"[{u['unit_type']} {u['unit']}]\n{u['text']}" for u in units if u.get("text"))
        EXTRACTED.mkdir(parents=True, exist_ok=True)
        text_path = EXTRACTED / f"{doc.id}.txt"
        text_path.write_text(text, encoding="utf-8")
        doc.text_path = str(text_path.relative_to(ROOT))
        doc.extracted_chars = len(text)
        doc.extracted_units = len(units)
        if not units:
            doc.error = (doc.error + " " if doc.error else "") + "no_text_extracted"
    except Exception as exc:
        doc.error = (doc.error + " " if doc.error else "") + f"extract_failed:{exc!r}"
        log(f"extract failed: {doc.title} -> {exc}")
    return doc, units


def split_text(text: str, max_chars: int = 1800, overlap: int = 250) -> list[str]:
    text = clean_text(text)
    if not text:
        return []
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    chunks: list[str] = []
    current = ""
    for para in paragraphs:
        if len(para) > max_chars:
            if current:
                chunks.append(current.strip())
                current = ""
            start = 0
            while start < len(para):
                chunks.append(para[start : start + max_chars].strip())
                start += max_chars - overlap
            continue
        if len(current) + len(para) + 2 <= max_chars:
            current = f"{current}\n\n{para}" if current else para
        else:
            if current:
                chunks.append(current.strip())
            tail = current[-overlap:] if current and overlap else ""
            current = f"{tail}\n\n{para}".strip() if tail else para
    if current:
        chunks.append(current.strip())
    return [c for c in chunks if len(c) > 40]


def build_chunks(docs: list[Document], extracted: dict[str, list[dict]]) -> list[dict]:
    chunks: list[dict] = []
    for doc in docs:
        if doc.status != "downloaded":
            continue
        units = extracted.get(doc.id, [])
        for unit in units:
            pieces = split_text(unit.get("text", ""))
            for idx, piece in enumerate(pieces):
                chunk_id = f"{doc.id}:{unit.get('unit_type')}:{unit.get('unit')}:{idx}"
                context_path = []
                for ctx in doc.contexts:
                    for item in ctx.get("path", []):
                        if item and item not in context_path:
                            context_path.append(item)
                chunks.append(
                    {
                        "chunk_id": chunk_id,
                        "doc_id": doc.id,
                        "title": doc.title,
                        "category": doc.category,
                        "family": doc.family,
                        "release": doc.release,
                        "revision_status": doc.revision_status,
                        "context_path": context_path,
                        "unit_type": unit.get("unit_type"),
                        "unit": unit.get("unit"),
                        "text": piece,
                        "local_path": doc.local_path,
                        "source_url": doc.url,
                    }
                )
    return chunks


def build_index(docs: list[Document], chunks: list[dict]) -> None:
    PROCESSED.mkdir(parents=True, exist_ok=True)
    docs_payload = [doc.to_dict() for doc in docs]
    (PROCESSED / "documents.json").write_text(json.dumps(docs_payload, indent=2, ensure_ascii=False), encoding="utf-8")
    with (PROCESSED / "documents.jsonl").open("w", encoding="utf-8") as fh:
        for doc in docs_payload:
            fh.write(json.dumps(doc, ensure_ascii=False) + "\n")
    with (PROCESSED / "chunks.jsonl").open("w", encoding="utf-8") as fh:
        for chunk in chunks:
            fh.write(json.dumps(chunk, ensure_ascii=False) + "\n")

    if not chunks:
        raise RuntimeError("No chunks were produced; cannot build index")

    search_texts = [
        "\n".join(
            [
                chunk.get("title", ""),
                chunk.get("family", ""),
                chunk.get("category", ""),
                chunk.get("release", ""),
                " > ".join(chunk.get("context_path", [])),
                chunk.get("text", ""),
            ]
        )
        for chunk in chunks
    ]
    word_vectorizer = TfidfVectorizer(
        lowercase=True,
        strip_accents="unicode",
        ngram_range=(1, 3),
        min_df=1,
        max_features=260000,
        sublinear_tf=True,
        token_pattern=r"(?u)\b[\w./-]{2,}\b",
    )
    char_vectorizer = TfidfVectorizer(
        lowercase=True,
        strip_accents="unicode",
        analyzer="char_wb",
        ngram_range=(3, 5),
        min_df=1,
        max_features=180000,
        sublinear_tf=True,
    )
    log(f"building word index over {len(chunks)} chunks")
    word_matrix = word_vectorizer.fit_transform(search_texts)
    log("building character index")
    char_matrix = char_vectorizer.fit_transform(search_texts)
    bm25_vectorizer = CountVectorizer(
        lowercase=True,
        strip_accents="unicode",
        ngram_range=(1, 2),
        min_df=1,
        max_features=220000,
        token_pattern=r"(?u)\b[\w./-]{2,}\b",
    )
    log("building BM25 lexical index")
    bm25_matrix = bm25_vectorizer.fit_transform(search_texts).tocsr()
    bm25_doc_len = np.asarray(bm25_matrix.sum(axis=1)).ravel().astype(float)
    bm25_avgdl = float(bm25_doc_len.mean() or 1.0)
    bm25_df = np.asarray((bm25_matrix > 0).sum(axis=0)).ravel().astype(float)
    bm25_idf = np.log(((len(chunks) - bm25_df + 0.5) / (bm25_df + 0.5)) + 1.0)
    payload = {
        "built_at": utc_now(),
        "index_url": INDEX_URL,
        "docs": docs_payload,
        "chunks": chunks,
        "chunk_id_to_pos": {chunk.get("chunk_id"): idx for idx, chunk in enumerate(chunks) if chunk.get("chunk_id")},
        "word_vectorizer": word_vectorizer,
        "word_matrix": word_matrix,
        "char_vectorizer": char_vectorizer,
        "char_matrix": char_matrix,
        "bm25_vectorizer": bm25_vectorizer,
        "bm25_matrix": bm25_matrix,
        "bm25_idf": bm25_idf,
        "bm25_doc_len": bm25_doc_len,
        "bm25_avgdl": bm25_avgdl,
    }
    with (PROCESSED / "index.pkl").open("wb") as fh:
        pickle.dump(payload, fh, protocol=pickle.HIGHEST_PROTOCOL)

    summary = {
        "built_at": payload["built_at"],
        "source": INDEX_URL,
        "documents_total": len(docs),
        "documents_downloaded": sum(1 for d in docs if d.status == "downloaded"),
        "documents_failed": sum(1 for d in docs if d.status == "failed"),
        "chunks": len(chunks),
        "index_flavour": "premium_hybrid_tfidf_char_bm25",
        "bm25_features": len(bm25_vectorizer.vocabulary_),
        "categories": sorted({d.category for d in docs}),
        "families": sorted({d.family for d in docs}),
    }
    (PROCESSED / "manifest.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")


def rebuild_index_from_processed() -> dict:
    """Rebuild vector indexes from processed JSON files.

    This is used by downstream ingesters, e.g. MyStandards, after appending
    extra local documents/chunks to the processed corpus.
    """
    documents_path = PROCESSED / "documents.json"
    chunks_path = PROCESSED / "chunks.jsonl"
    if not documents_path.exists() or not chunks_path.exists():
        raise FileNotFoundError("Processed documents/chunks not found; run tips_ingest.py first")
    docs_payload = json.loads(documents_path.read_text(encoding="utf-8"))
    docs = [Document(**doc) for doc in docs_payload]
    chunks: list[dict] = []
    with chunks_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                chunks.append(json.loads(line))
    build_index(docs, chunks)
    return json.loads((PROCESSED / "manifest.json").read_text(encoding="utf-8"))


def write_catalog(docs: list[Document], skipped: list[dict]) -> None:
    lines = ["# TIPS local catalog", "", f"Generated: {utc_now()}", "", "## Documents", ""]
    for doc in sorted(docs, key=lambda d: (d.category, d.release, d.family, d.title.lower())):
        status = doc.status
        release = f" | {doc.release}" if doc.release else ""
        local = doc.local_path or ""
        lines.append(f"- [{doc.category}/{doc.family}{release}] {doc.title} ({status})")
        lines.append(f"  - Local: `{local}`")
        lines.append(f"  - Source: {doc.url}")
    if skipped:
        lines.extend(["", "## Skipped external/media links", ""])
        for item in skipped:
            lines.append(f"- {item.get('title')} - {item.get('reason')} - {item.get('url')}")
    (PROCESSED / "catalog.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    (PROCESSED / "skipped_links.json").write_text(json.dumps(skipped, indent=2, ensure_ascii=False), encoding="utf-8")


def ingest(force: bool = False, include_media: bool = False, limit: int | None = None) -> dict:
    for folder in [RAW_DOCS, EXTRACTED, PROCESSED]:
        folder.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT, "Accept": "*/*"})
    html = fetch_index(session, force=force)
    docs, skipped = parse_index(html, include_media=include_media)
    if limit:
        docs = docs[:limit]
    log(f"found {len(docs)} unique local-document candidates; skipped {len(skipped)} external/media links")

    downloaded: list[Document] = []
    for idx, doc in enumerate(docs, start=1):
        log(f"downloading {idx}/{len(docs)}: {doc.title[:90]}")
        downloaded.append(download_doc(session, doc, force=force))
        time.sleep(0.05)

    extracted: dict[str, list[dict]] = {}
    for idx, doc in enumerate(downloaded, start=1):
        if doc.status != "downloaded":
            continue
        log(f"extracting {idx}/{len(downloaded)}: {doc.title[:90]}")
        doc, units = extract_text_for_doc(doc)
        extracted[doc.id] = units

    chunks = build_chunks(downloaded, extracted)
    log(f"built {len(chunks)} retrieval chunks")
    build_index(downloaded, chunks)
    write_catalog(downloaded, skipped)

    manifest = json.loads((PROCESSED / "manifest.json").read_text(encoding="utf-8"))
    log(f"done: {manifest['documents_downloaded']} docs, {manifest['chunks']} chunks")
    return manifest


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Ingest ECB TIPS documents into a local optimized index.")
    parser.add_argument("--force", action="store_true", help="redownload and rebuild from scratch")
    parser.add_argument("--include-media", action="store_true", help="also download linked media files such as MP4")
    parser.add_argument("--limit", type=int, default=None, help="debug: ingest only first N documents")
    args = parser.parse_args(argv)
    try:
        ingest(force=args.force, include_media=args.include_media, limit=args.limit)
        return 0
    except KeyboardInterrupt:
        return 130
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
